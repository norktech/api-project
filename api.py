from fastapi import FastAPI
import openpyxl
from datetime import datetime

app = FastAPI()

@app.get("/")
def home():
    return {"mensagem": "Bem vinda à NorkTech API!", "status": "online"}

@app.get("/livros")
def listar_livros():
    wb = openpyxl.load_workbook("produtos.xlsx")
    ws = wb.active
    livros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        livros.append({
            "nome": row[0],
            "preco": row[1],
            "data": row[2]
        })
    return {"total": len(livros), "livros": livros}

@app.get("/livros/{nome}")
def buscar_livro(nome: str):
    wb = openpyxl.load_workbook("produtos.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if nome.lower() in row[0].lower():
            return {"nome": row[0], "preco": row[1], "data": row[2]}
    return {"erro": "Livro não encontrado"}